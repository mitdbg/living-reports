#!/usr/bin/env python3

import json
import os
import logging
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
from openai import OpenAI

# Make together import optional
try:
    from together import Together
except ImportError:
    Together = None

from chat_manager import ChatManager
from template import Template
from execution_result import ExecutionResult
from simple_view import SimpleView
from diff_view import DiffView  
from translate_comments_to_code_change import CommentTranslationService, CommentContext

# Add file processing imports
import pandas as pd
from bs4 import BeautifulSoup
import PyPDF2
import openpyxl
from pathlib import Path
import io
import base64
from PIL import Image
import tempfile
import os

# Add parent directory to path for imports
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger('backend')

app = Flask(__name__)
CORS(app)

# Initialize OpenAI client
try:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        logger.warning("OPENAI_API_KEY not set! AI features will be limited.")
        logger.info("Set it with: export OPENAI_API_KEY='your-key'")
    
    client = OpenAI(api_key=api_key) if api_key else None
    if client:
        logger.info("✓ OpenAI client initialized successfully")
    else:
        api_key = os.getenv("TOGETHER_API_KEY")
        if api_key and Together:
            client = Together(api_key=api_key)
            logger.info("✓ Together client initialized successfully")
        else:
            if not Together:
                logger.warning("⚠ Together module not available (install with: pip install together)")
            logger.warning("⚠ Running without OpenAI or Together client")
        
except Exception as e:
    logger.error(f"Failed to initialize OpenAI client: {e}")
    client = None

# Global state
view_registry = {}  # session_id -> View
chat_manager = ChatManager(client, view_registry) if client else None

# Initialize comment translation service
comment_translation_service = CommentTranslationService(client) if client else None

# Persistent storage for all documents
DATABASE_DIR = 'database'
DOCUMENTS_FILE = os.path.join(DATABASE_DIR, 'documents.json')

def ensure_database_dir():
    """Ensure database directory exists"""
    if not os.path.exists(DATABASE_DIR):
        os.makedirs(DATABASE_DIR)
        logger.info(f"📁 Created database directory: {DATABASE_DIR}")

def load_documents():
    """Load all documents from file on startup"""
    global documents
    try:
        ensure_database_dir()
        if os.path.exists(DOCUMENTS_FILE):
            with open(DOCUMENTS_FILE, 'r') as f:
                documents = json.load(f)
                logger.info(f"📄 Loaded {len(documents)} documents from {DOCUMENTS_FILE}")
        else:
            documents = {}
            logger.info("📄 No existing documents file found. Starting fresh.")
    except Exception as e:
        logger.error(f"❌ Error loading documents: {e}")
        documents = {}

def save_documents():
    """Save all documents to file for persistence"""
    try:
        ensure_database_dir()
        with open(DOCUMENTS_FILE, 'w') as f:
            json.dump(documents, f, indent=2)
        logger.info(f"💾 Saved {len(documents)} documents to {DOCUMENTS_FILE}")
    except Exception as e:
        logger.error(f"❌ Error saving documents: {e}")

# Initialize storage
documents = {}  # Global storage for all documents
load_documents()

# Persistent storage for document verifications
VERIFICATIONS_FILE = os.path.join(DATABASE_DIR, 'verifications.json')

def load_verifications():
    """Load all verifications from file"""
    try:
        ensure_database_dir()
        if os.path.exists(VERIFICATIONS_FILE):
            with open(VERIFICATIONS_FILE, 'r') as f:
                verifications = json.load(f)
                logger.info(f"📋 Loaded verifications for {len(verifications)} documents from {VERIFICATIONS_FILE}")
                return verifications
        else:
            logger.info("📋 No existing verifications file found. Starting fresh.")
            return {}
    except Exception as e:
        logger.error(f"❌ Error loading verifications: {e}")
        return {}

def save_verifications(verifications):
    """Save all verifications to file"""
    try:
        ensure_database_dir()
        with open(VERIFICATIONS_FILE, 'w') as f:
            json.dump(verifications, f, indent=2)
        logger.info(f"💾 Saved verifications for {len(verifications)} documents to {VERIFICATIONS_FILE}")
    except Exception as e:
        logger.error(f"❌ Error saving verifications: {e}")

# Initialize verifications storage
verifications = load_verifications()

# Persistent storage for Data Lake
DATA_LAKE_FILE = os.path.join(DATABASE_DIR, 'data_lake.json')

def load_data_lake():
    """Load all data lake items from file"""
    try:
        ensure_database_dir()
        if os.path.exists(DATA_LAKE_FILE):
            with open(DATA_LAKE_FILE, 'r') as f:
                data_lake = json.load(f)
                logger.info(f"🗂️ Loaded data lake for {len(data_lake)} documents from {DATA_LAKE_FILE}")
                return data_lake
        else:
            logger.info("🗂️ No existing data lake file found. Starting fresh.")
            return {}
    except Exception as e:
        logger.error(f"❌ Error loading data lake: {e}")
        return {}

def save_data_lake(data_lake):
    """Save all data lake items to file"""
    try:
        ensure_database_dir()
        with open(DATA_LAKE_FILE, 'w') as f:
            json.dump(data_lake, f, indent=2)
        logger.info(f"💾 Saved data lake for {len(data_lake)} documents to {DATA_LAKE_FILE}")
    except Exception as e:
        logger.error(f"❌ Error saving data lake: {e}")

# Initialize data lake storage
data_lake_storage = load_data_lake()

# Persistent storage for Variables
VARIABLES_FILE = os.path.join(DATABASE_DIR, 'vars.json')

def load_variables():
    """Load all variables from file"""
    try:
        ensure_database_dir()
        if os.path.exists(VARIABLES_FILE):
            with open(VARIABLES_FILE, 'r') as f:
                variables = json.load(f)
                logger.info(f"📊 Loaded variables for {len(variables)} documents from {VARIABLES_FILE}")
                return variables
        else:
            logger.info("📊 No existing variables file found. Starting fresh.")
            return {}
    except Exception as e:
        logger.error(f"❌ Error loading variables: {e}")
        return {}

def save_variables(variables):
    """Save all variables to file"""
    try:
        ensure_database_dir()
        with open(VARIABLES_FILE, 'w') as f:
            json.dump(variables, f, indent=2)
        logger.info(f"💾 Saved variables for {len(variables)} documents to {VARIABLES_FILE}")
    except Exception as e:
        logger.error(f"❌ Error saving variables: {e}")

# Initialize variables storage
variables_storage = load_variables()

@app.before_request
def log_request():
    """Log all incoming requests for debugging."""
    # print(f"📥 Incoming request: {request.method} {request.url}", flush=True)
    if request.method == 'POST' and request.is_json:
        print(f"📄 Request data: {request.get_json()}", flush=True)

@app.route('/api/chat', methods=['POST'])
def handle_chat():
    """Handle chat messages using the ChatManager."""
    try:
        data = request.get_json()
        user_message = data.get('message', '')
        session_id = data.get('session_id', 'default')
        current_template = data.get('current_template', '')
        suggest_template = data.get('suggest_template', False)  # Default to False for regular chat
        
        # Check if chat_manager is available (requires OpenAI API key)
        if chat_manager is None:
            return jsonify({
                'error': 'OpenAI API key not configured',
                'response': 'Sorry, I cannot process chat messages because the OpenAI API key is not set. Please set the OPENAI_API_KEY environment variable and restart the server.',
                'content': 'To enable AI chat features, please:\n1. Get an OpenAI API key from https://platform.openai.com/\n2. Set it as an environment variable: export OPENAI_API_KEY="your-key-here"\n3. Restart the Python backend',
                'result': '',
                'variables': {},
                'cache_info': {'variables_count': 0, 'cached': False},
                'template': {'current_template': current_template, 'view_type': 'simple'},
                'output': {'result': '', 'variables': {}, 'view_type': 'simple'},
                'view_type': 'simple'
            })
        
        # Use chat_manager to handle the message
        response = chat_manager.handle_chat_message(
            user_message=user_message,
            session_id=session_id,
            current_template_text=current_template,
            suggest_template=suggest_template
        )
        
        return jsonify(response)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'content': f'Error processing message: {str(e)}'
        }), 500

@app.route('/api/execute-template', methods=['POST'])
def execute_template():
    """Execute a template and return the result."""
    try:
        data = request.get_json()
        template_text = data.get('template_text', '')
        session_id = data.get('session_id', 'default')
        document_id = data.get('document_id', None)
        
        # Create or get the view for this session
        if session_id not in view_registry:
            template = Template(template_text, document_id)
            execution_result = ExecutionResult()
            view_registry[session_id] = SimpleView(template, execution_result, client)
        
        # Update the template and execute it
        view = view_registry[session_id]
        view.update_from_editor(template_text, document_id)
        
        # Get the rendered output
        output_data = view.render_output()
        template_data = view.render_template()
        
        print(output_data, flush=True)
        return jsonify({
            'success': True,
            'template_text': template_data.get('template_text', template_text),
            'rendered_output': output_data.get('result', ''),
            'variables': output_data.get('variables', {}),
            'view_type': output_data.get('view_type', 'simple')
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'rendered_output': f'Error executing template: {str(e)}'
        }), 500

@app.route('/api/file-context', methods=['POST'])
def set_file_context():
    """Set file context for chat."""
    try:
        data = request.get_json()
        file_name = data.get('fileName', '')
        content = data.get('content', '')
        session_id = data.get('session_id', 'default')
        
        # Create a template from the file content
        template = Template(content)
        execution_result = ExecutionResult()
        view_registry[session_id] = SimpleView(template, execution_result, client)
        
        return jsonify({
            'message': f'File "{file_name}" has been loaded as context for chat.',
            'success': True
        })
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'success': False
        }), 500

@app.route('/api/file-context', methods=['DELETE'])
def clear_file_context():
    """Clear file context for chat."""
    try:
        session_id = request.args.get('session_id', 'default')
        
        if session_id in view_registry:
            del view_registry[session_id]
        
        return jsonify({
            'message': 'File context has been cleared.',
            'success': True
        })
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'success': False
        }), 500

@app.route('/api/translate-comment', methods=['POST'])
def translate_comment():
    """Translate user comment into template edit suggestion."""
    try:
        data = request.get_json()
        
        # Extract comment data
        comment_text = data.get('comment_text', '')
        selected_text = data.get('selected_text', '')
        mode = data.get('mode', 'preview')  # 'preview', 'template', 'source'
        
        # Extract document context
        template_content = data.get('template_content', '')
        preview_content = data.get('preview_content', '')
        source_content = data.get('source_content', '')
        variables = data.get('variables', {})
        
        # Document and session info
        document_id = data.get('document_id')
        session_id = data.get('session_id', 'default')
        
        # Validate required fields
        if not comment_text:
            return jsonify({
                'success': False,
                'error': 'Comment text is required'
            }), 400
        
        # Check if comment translation service is available
        if comment_translation_service is None:
            return jsonify({
                'success': False,
                'error': 'Comment translation service not available (API key not configured)',
                'fallback_suggestion': {
                    'original_comment': comment_text,
                    'selected_text': selected_text,
                    'suggested_change': f"Review and address the comment: '{comment_text}'",
                    'explanation': "Please manually review this comment and update the template accordingly. AI assistance is not available without an API key.",
                    'confidence': 0.5,
                    'change_type': 'manual_review'
                }
            })
        
        # Create comment context
        comment_context = CommentContext(
            comment_text=comment_text,
            selected_text=selected_text,
            mode=mode,
            template_content=template_content,
            preview_content=preview_content,
            source_content=source_content,
            variables=variables
        )
        
        # Get translation suggestion
        suggestion = comment_translation_service.translate_comment_to_template_edit(comment_context)
        
        print(f"_____Suggestion: {suggestion}")
        # Convert suggestion to dict for JSON response
        suggestion_dict = {
            'original_comment': suggestion.original_comment,
            'selected_text': suggestion.selected_text,
            'suggested_change': suggestion.suggested_change,
            'explanation': suggestion.explanation,
            'confidence': suggestion.confidence,
            'change_type': suggestion.change_type,
            'original_text': suggestion.original_text
        }
        
        logger.info(f"Generated comment translation for document {document_id}: {comment_text[:50]}...")
        
        return jsonify({
            'success': True,
            'suggestion': suggestion_dict,
            'analysis': {
                'comment_length': len(comment_text),
                'selected_text_length': len(selected_text),
                'mode': mode,
                'template_length': len(template_content),
                'variables_count': len(variables)
            }
        })
        
    except Exception as e:
        logger.error(f"Error translating comment: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'fallback_suggestion': {
                'original_comment': comment_text if 'comment_text' in locals() else '',
                'selected_text': selected_text if 'selected_text' in locals() else '',
                'suggested_change': f"Error occurred while processing comment. Please review manually.",
                'explanation': f"An error occurred: {str(e)}",
                'confidence': 0.1,
                'change_type': 'error'
            }
        }), 500

@app.route('/api/ai-suggestion', methods=['POST'])
def get_ai_suggestion():
    """Get AI suggestion for content improvement in specific mode (preview, template, source)."""
    try:
        data = request.get_json()
        
        # Extract request data
        full_content = data.get('full_content', '')
        selected_text = data.get('selected_text', '')
        user_request = data.get('user_request', '')
        mode = data.get('mode', 'preview')  # 'preview', 'template', 'source'
        session_id = data.get('session_id', 'default')
        
        # Validate required fields
        if not full_content:
            return jsonify({
                'success': False,
                'error': 'Full content is required'
            }), 400
            
        if not selected_text:
            return jsonify({
                'success': False,
                'error': 'Selected text is required'
            }), 400
            
        if not user_request:
            return jsonify({
                'success': False,
                'error': 'User request is required'
            }), 400
        
        # Check if LLM client is available
        if client is None:
            return jsonify({
                'success': False,
                'error': 'AI service not available (API key not configured)',
                'fallback_suggestion': {
                    'change_type': 'manual_review',
                    'original_text': selected_text,
                    'new_text': f'Please review: {user_request}',
                    'explanation': 'AI assistance is not available without an API key.',
                    'confidence': 0.1
                }
            })
        
        # Create structured prompt for LLM
        prompt = f"""You are an AI assistant helping to improve content based on user feedback.

MODE: {mode}
FULL CONTENT:
{full_content}

SELECTED TEXT: "{selected_text}"
USER REQUEST: "{user_request}"

Based on the user's request, please suggest a specific improvement to the selected text.

IMPORTANT: Respond with ONLY a JSON object in this exact format:
{{
    "change_type": "replace|add|remove",
    "original_text": "exact text from content to change",
    "new_text": "suggested replacement/addition text",
    "explanation": "brief explanation of the change",
    "confidence": 0.85
}}

Requirements:
1. "original_text" must be the exact text that exists in the content
2. For "replace": provide both original_text and new_text
3. For "add": original_text can be nearby context, new_text is what to add
4. For "remove": only original_text is needed, new_text can be empty
5. "confidence" should be between 0.0 and 1.0
6. Return ONLY the JSON object, no other text

JSON:"""
        
        try:
            # Call LLM for suggestion
            if hasattr(client, 'chat'):
                # OpenAI client
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.3,
                    max_tokens=500
                )
                suggestion_text = response.choices[0].message.content.strip()
            else:
                # Together client
                response = client.chat.completions.create(
                    model="Qwen/Qwen2.5-Coder-32B-Instruct",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.3,
                    max_tokens=500
                )
                suggestion_text = response.choices[0].message.content.strip()
            
            print(f"AI Suggestion Raw Response: {suggestion_text}")
            
            # Parse LLM response
            try:
                # Try to extract JSON from the response
                import json
                import re
                
                # Clean up the response to extract JSON
                json_match = re.search(r'\{[\s\S]*\}', suggestion_text)
                if json_match:
                    json_str = json_match.group()
                    parsed_suggestion = json.loads(json_str)
                else:
                    raise ValueError("No JSON found in response")
                
                # Validate required fields
                required_fields = ['change_type', 'original_text', 'explanation']
                for field in required_fields:
                    if field not in parsed_suggestion:
                        raise ValueError(f"Missing required field: {field}")
                
                # Validate change_type
                if parsed_suggestion['change_type'] not in ['replace', 'add', 'remove']:
                    raise ValueError(f"Invalid change_type: {parsed_suggestion['change_type']}")
                
                # Set defaults
                parsed_suggestion['confidence'] = parsed_suggestion.get('confidence', 0.7)
                parsed_suggestion['new_text'] = parsed_suggestion.get('new_text', '')
                
                # Additional validation: check if original_text exists in full_content
                if parsed_suggestion['original_text'] and parsed_suggestion['original_text'] not in full_content:
                    print(f"Warning: original_text '{parsed_suggestion['original_text']}' not found in content")
                    # Try to find similar text
                    if selected_text in full_content:
                        parsed_suggestion['original_text'] = selected_text
                        print(f"Using selected_text as fallback: '{selected_text}'")
                
                return jsonify({
                    'success': True,
                    'suggestion': parsed_suggestion,
                    'mode': mode,
                    'raw_response': suggestion_text
                })
                
            except (json.JSONDecodeError, ValueError) as parse_error:
                print(f"Error parsing AI response: {parse_error}")
                # Fallback: create a simple suggestion
                fallback_suggestion = {
                    'change_type': 'replace',
                    'original_text': selected_text,
                    'new_text': suggestion_text[:200],  # Use first part of response
                    'explanation': f'AI suggested improvement: {suggestion_text[:100]}...',
                    'confidence': 0.6
                }
                
                return jsonify({
                    'success': True,
                    'suggestion': fallback_suggestion,
                    'mode': mode,
                    'raw_response': suggestion_text,
                    'fallback': True
                })
                
        except Exception as llm_error:
            print(f"Error calling LLM: {llm_error}")
            return jsonify({
                'success': False,
                'error': f'LLM error: {str(llm_error)}',
                'fallback_suggestion': {
                    'change_type': 'replace',
                    'original_text': selected_text,
                    'new_text': f'Consider: {user_request}',
                    'explanation': f'Failed to get AI suggestion: {str(llm_error)}',
                    'confidence': 0.3
                }
            }), 500
        
    except Exception as e:
        print(f"Error in AI suggestion endpoint: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/compute-diff', methods=['POST'])
def compute_diff():
    """Compute diff between current and suggested template/output."""
    try:
        data = request.get_json()
        current_text = data.get('current_text', '')
        suggested_text = data.get('suggested_text', '')
        session_id = data.get('session_id', 'default')
        content_type = data.get('content_type', 'template')  # 'template' or 'output'
        
        # Use the backend diff computation from DiffViewStrategy
        if content_type == 'template':
            # For templates, we need to execute both to get outputs
            document_id = data.get('document_id', None)
            current_template = Template(current_text, document_id)
            suggested_template = Template(suggested_text, document_id)
            
            # Execute both templates
            execution_result = ExecutionResult()
            current_result = current_template.execute(client, execution_result)
            suggested_result = suggested_template.execute(client, execution_result)
            
            # Create DiffView to compute diffs
            from diff_view import DiffView
            diff_view = DiffView(
                current_template=current_template,
                current_result=current_result,
                suggested_template=suggested_template,
                suggested_result=suggested_result,
                client=client
            )
            
            template_data = diff_view.render_template()
            output_data = diff_view.render_output()
            
            return jsonify({
                'success': True,
                'template_diffs': template_data.get('line_diffs', []),
                'current_template': template_data.get('current_template', ''),
                'suggested_template': template_data.get('suggested_template', ''),
                'current_output': output_data.get('current_output', ''),
                'suggested_output': output_data.get('suggested_output', ''),
                'output_diffs': compute_text_diff(
                    output_data.get('current_output', ''),
                    output_data.get('suggested_output', '')
                )
            })
        else:
            # For output-only diff
            diffs = compute_text_diff(current_text, suggested_text)
            return jsonify({
                'success': True,
                'output_diffs': diffs,
                'current_output': current_text,
                'suggested_output': suggested_text
            })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def compute_text_diff(current_text, suggested_text):
    """Helper function to compute line-by-line diff for any text."""
    current_lines = current_text.split('\n')
    suggested_lines = suggested_text.split('\n')
    max_lines = max(len(current_lines), len(suggested_lines))
    
    diffs = []
    for i in range(max_lines):
        current_line = current_lines[i] if i < len(current_lines) else ''
        suggested_line = suggested_lines[i] if i < len(suggested_lines) else ''
        
        if current_line != suggested_line:
            diffs.append({
                'line_index': i,
                'current_line': current_line,
                'suggested_line': suggested_line,
                'change_type': 'modified' if current_line and suggested_line 
                             else 'added' if not current_line 
                             else 'removed'
            })
    
    return diffs

@app.route('/api/view-action', methods=['POST'])
def handle_view_action():
    """Handle view actions like accept/reject in diff view."""
    try:
        data = request.get_json()
        action = data.get('action', '')
        session_id = data.get('session_id', 'default')
        
        if session_id not in view_registry:
            return jsonify({'error': 'No view found for session'}), 404
        
        view = view_registry[session_id]
        
        if isinstance(view, DiffView):
            if action == 'accept':
                new_view = view.accept_suggestion()
                view_registry[session_id] = new_view
                return jsonify({
                    'success': True,
                    'message': 'Suggestion accepted',
                    'view_type': 'simple'
                })
            elif action == 'reject':
                new_view = view.reject_suggestion()
                view_registry[session_id] = new_view
                return jsonify({
                    'success': True,
                    'message': 'Suggestion rejected',
                    'view_type': 'simple'
                })
        
        return jsonify({'error': 'Invalid action or view type'}), 400
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'success': False
        }), 500

@app.route('/api/get-view', methods=['GET'])
def get_current_view():
    """Get the current view state."""
    try:
        session_id = request.args.get('session_id', 'default')
        
        if session_id not in view_registry:
            return jsonify({
                'view_type': 'simple',
                'template_text': '',
                'rendered_output': '',
                'variables': {}
            })
        
        view = view_registry[session_id]
        template_data = view.render_template()
        output_data = view.render_output()
        
        response = {
            'view_type': output_data.get('view_type', 'simple'),
            'template_text': template_data.get('template_text', ''),
            'rendered_output': output_data.get('result', ''),
            'variables': output_data.get('variables', {})
        }
        
        # Add diff-specific data if it's a DiffView
        if isinstance(view, DiffView):
            response.update({
                'current_template': template_data.get('current_template', ''),
                'suggested_template': template_data.get('suggested_template', ''),
                'line_diffs': template_data.get('line_diffs', []),
                'current_output': output_data.get('current_output', ''),
                'suggested_output': output_data.get('suggested_output', '')
            })
        
        return jsonify(response)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'view_type': 'simple',
            'template_text': '',
            'rendered_output': '',
            'variables': {}
        }), 500

@app.route('/api/chat/clear', methods=['POST'])
def clear_chat_history():
    """Clear chat history for a session."""
    try:
        data = request.get_json()
        session_id = data.get('session_id', 'default')
        
        # Clear conversation history in chat manager if it exists
        cleared = False
        if chat_manager:
            cleared = chat_manager.clear_conversation_history(session_id)
        
        return jsonify({
            'success': True,
            'cleared': cleared,
            'message': f'Chat history cleared for session {session_id}' if cleared else f'No chat history found for session {session_id}'
        })
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'success': False
        }), 500

# File processing utility functions
def process_excel_file(file_content, file_name):
    """Process Excel file and extract text content."""
    try:
        # Create a BytesIO object from the content
        file_buffer = io.BytesIO(base64.b64decode(file_content))
        
        # Read Excel file
        if file_name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_buffer)
            content = f"Excel File: {file_name}\n\n"
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content += f"Sheet: {sheet_name}\n"
                content += "-" * 40 + "\n"
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                        content += row_text + "\n"
                content += "\n"
        else:
            # For .xls files, use pandas
            df_dict = pd.read_excel(file_buffer, sheet_name=None)
            content = f"Excel File: {file_name}\n\n"
            
            for sheet_name, df in df_dict.items():
                content += f"Sheet: {sheet_name}\n"
                content += "-" * 40 + "\n"
                content += df.to_string(index=False) + "\n\n"
        
        return content
        
    except Exception as e:
        return f"Error processing Excel file: {str(e)}"

def process_pdf_file(file_content, file_name):
    """Process PDF file and extract text content."""
    try:
        # Create a BytesIO object from the content
        file_buffer = io.BytesIO(base64.b64decode(file_content))
        
        # Read PDF file
        pdf_reader = PyPDF2.PdfReader(file_buffer)
        content = f"PDF File: {file_name}\n"
        content += f"Number of pages: {len(pdf_reader.pages)}\n\n"
        
        for page_num, page in enumerate(pdf_reader.pages, 1):
            content += f"Page {page_num}:\n"
            content += "-" * 40 + "\n"
            try:
                page_text = page.extract_text()
                content += page_text + "\n\n"
            except Exception as e:
                content += f"Error extracting text from page {page_num}: {str(e)}\n\n"
        
        return content
        
    except Exception as e:
        return f"Error processing PDF file: {str(e)}"

def process_html_file(file_content, file_name):
    """Process HTML file and extract text content."""
    try:
        # If content is base64 encoded, decode it
        if isinstance(file_content, str) and len(file_content) > 100:
            try:
                html_content = base64.b64decode(file_content).decode('utf-8')
            except:
                html_content = file_content
        else:
            html_content = file_content
        
        # Parse HTML and extract text
        soup = BeautifulSoup(html_content, 'html.parser')
        
        content = f"HTML File: {file_name}\n\n"
        
        # Extract title if available
        title = soup.find('title')
        if title:
            content += f"Title: {title.get_text().strip()}\n\n"
        
        # Extract main content
        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()
        
        # Get text content
        text = soup.get_text()
        
        # Clean up whitespace
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        content += '\n'.join(chunk for chunk in chunks if chunk)
        
        return content
        
    except Exception as e:
        return f"Error processing HTML file: {str(e)}"

def process_pptx_file(file_path):
    """
    PPTX files are now processed client-side using PPTX2HTML JavaScript library.
    This function returns a message indicating the file should be processed in the browser.
    """
    try:
        # Get basic file info
        file_size = os.path.getsize(file_path)
        
        return {
            'success': True,
            'content': f'PPTX file ready for client-side processing. File size: {file_size} bytes. This file will be processed using the PPTX2HTML JavaScript library in your browser for better visual fidelity.',
            'file_type': 'pptx',
            'processing_method': 'client-side'
        }
    except Exception as e:
        return {
            'success': False,
            'error': f'Error accessing PPTX file: {str(e)}'
        }

@app.route('/api/process-file', methods=['POST'])
def process_file():
    """Process uploaded files (Excel, PDF, HTML) and return extracted content."""
    try:
        data = request.get_json()
        file_name = data.get('fileName', '')
        file_content = data.get('content', '')  # Base64 encoded for binary files
        file_path = data.get('filePath', '')
        session_id = data.get('session_id', 'default')
        
        # Determine file type and process accordingly
        file_ext = Path(file_name).suffix.lower()
        
        if file_ext in ['.xlsx', '.xls']:
            processed_content = process_excel_file(file_content, file_name)
        elif file_ext == '.pdf':
            processed_content = process_pdf_file(file_content, file_name)
        elif file_ext in ['.html', '.htm']:
            processed_content = process_html_file(file_content, file_name)
        elif file_ext in ['.pptx', '.ppt']:
            processed_content = process_pptx_file(file_path)
        else:
            # For other file types, return as-is (assuming text)
            processed_content = file_content
        
        return jsonify({'success': True, 'message': 'File processed successfully', 'content': processed_content, 'fileName': file_name, 'filePath': file_path})
        
    except Exception as e:
        logger.error(f"Error processing file: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/documents', methods=['POST'])
def save_document():
    """Save a document to backend storage."""
    try:
        data = request.get_json()
        
        # Extract document data
        document_id = data.get('documentId')
        title = data.get('title', 'Untitled Document')
        template_content = data.get('template_content', '')
        source_content = data.get('source_content', '')
        preview_content = data.get('preview_content', '')
        session_id = data.get('sessionId')
        created_at = data.get('createdAt')
        last_modified = data.get('lastModified')
        chat_history = data.get('chatHistory', [])
        variables = data.get('variables', {})
        context_files = data.get('contextFiles', [])
        
        # Core document schema
        author = data.get('author')
        author_name = data.get('authorName', 'Unknown')
        editors = data.get('editors', [])
        viewers = data.get('viewers', [])
        
        # Extract comments data (new field)
        comments = data.get('comments', {})
        
        # Create document with unified schema
        document = {
            'id': document_id,
            'title': title,
            'source_content': source_content,
            'template_content': template_content,
            'preview_content': preview_content,
            'sessionId': session_id,
            'createdAt': created_at,
            'lastModified': last_modified,
            'chatHistory': chat_history,
            'variables': variables,
            'contextFiles': context_files,
            'author': author,
            'authorName': author_name,
            'editors': editors,
            'viewers': viewers,
            'comments': comments,  # Add comments to document schema
            'savedAt': datetime.now().isoformat()
        }
        
        # Store document
        documents[document_id] = document
        
        # Persist to file
        save_documents()
        
        # Log comment information for debugging
        comment_count = len(comments) if isinstance(comments, dict) else 0
        logger.info(f"Document '{title}' (ID: {document_id}) saved by {author_name} with {comment_count} comments")
        
        return jsonify({
            'success': True,
            'message': f'Document "{title}" has been saved',
            'documentId': document_id
        })
            
    except Exception as e:
        logger.error(f"Error saving document: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/documents', methods=['GET'])
def get_all_documents():
    """Get all documents."""
    try:
        documents_list = list(documents.values())
        
        logger.info(f"Returning {len(documents_list)} documents")
        
        return jsonify({
            'success': True,
            'documents': documents_list,
            'count': len(documents_list)
        })
        
    except Exception as e:
        logger.error(f"Error getting documents: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/documents/user/<user_id>', methods=['GET'])
def get_user_documents(user_id):
    """Get all documents accessible to a specific user."""
    try:
        accessible_docs = []
        
        for doc_id, doc in documents.items():
            # User can access if they are:
            # 1. The author
            # 2. In the editors list  
            # 3. In the viewers list
            if (doc.get('author') == user_id or
                user_id in doc.get('editors', []) or
                user_id in doc.get('viewers', [])):
                
                # Add permission info for frontend
                doc_copy = doc.copy()
                doc_copy['userPermission'] = {
                    'isAuthor': doc.get('author') == user_id,
                    'canEdit': doc.get('author') == user_id or user_id in doc.get('editors', []),
                    'canView': True  # If they can access it, they can view it
                }
                accessible_docs.append(doc_copy)
        
        # logger.info(f"Returning {len(accessible_docs)} accessible documents for user {user_id}")
        
        return jsonify({
            'success': True,
            'documents': accessible_docs,
            'count': len(accessible_docs)
        })
        
    except Exception as e:
        logger.error(f"Error getting documents for user {user_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/documents/<document_id>', methods=['GET'])
def get_shared_document(document_id):
    """Get a specific shared document by ID."""
    try:
        if document_id in documents:
            document = documents[document_id]
            # logger.info(f"Returning shared document: {document['title']}")
            
            return jsonify({
                'success': True,
                'document': document
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Document not found'
            }), 404
            
    except Exception as e:
        logger.error(f"Error getting shared document {document_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/documents/<document_id>', methods=['DELETE'])
def delete_document(document_id):
    """Delete a document."""
    try:
        if document_id in documents:
            document_title = documents[document_id]['title']
            del documents[document_id]
            
            # Persist changes
            save_documents()
            
            logger.info(f"Deleted document: {document_title}")
            
            return jsonify({
                'success': True,
                'message': f'Document "{document_title}" has been deleted'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Document not found'
            }), 404
            
    except Exception as e:
        logger.error(f"Error deleting document {document_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/verify-document', methods=['POST'])
def verify_document():
    """Save document verification."""
    try:
        data = request.get_json()
        
        # Extract verification data
        session_id = data.get('session_id')
        user_id = data.get('user_id')
        user_name = data.get('user_name')
        user_emoji = data.get('user_emoji')
        verified_at = data.get('verified_at')
        document_content = data.get('document_content', '')
        
        if not session_id or not user_id or not user_name:
            return jsonify({
                'success': False,
                'error': 'Missing required fields: session_id, user_id, user_name'
            }), 400
        
        # Create verification record
        verification = {
            'user_id': user_id,
            'user_name': user_name,
            'user_emoji': user_emoji,
            'verified_at': verified_at,
            'document_content_hash': hash(document_content) if document_content else None,
            'content_length': len(document_content) if document_content else 0,
            'saved_at': datetime.now().isoformat()
        }
        
        # Initialize verifications structure for this session if not exists
        if session_id not in verifications:
            verifications[session_id] = {}
        
        # Initialize user's verification list if not exists
        if user_id not in verifications[session_id]:
            verifications[session_id][user_id] = []
        
        # Add verification to the user's list
        verifications[session_id][user_id].append(verification)
        
        # Save to file
        save_verifications(verifications)
        
        logger.info(f"✅ Document verification saved: {user_name} verified document {session_id}")
        
        return jsonify({
            'success': True,
            'message': f'Document verified by {user_name}',
            'verification': verification
        })
        
    except Exception as e:
        logger.error(f"Error saving verification: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/get-verification/<session_id>', methods=['GET'])
def get_verification(session_id):
    """Get verification history for a document."""
    try:
        document_verifications = verifications.get(session_id, {})
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'verifications': document_verifications,
            'count': len(document_verifications)
        })
        
    except Exception as e:
        logger.error(f"Error getting verification for {session_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/data-lake', methods=['GET'])
def get_data_lake():
    """Get data lake items for a specific document."""
    try:
        document_id = request.args.get('documentId')
        window_id = request.args.get('windowId', 'default')
        session_id = request.args.get('session_id', 'default')
        
        if not document_id:
            return jsonify({
                'success': False,
                'error': 'Missing documentId parameter'
            }), 400
        
        # Get data lake items for this document
        document_data_lake = data_lake_storage.get(document_id, [])
        
        logger.info(f"🗂️ Returning {len(document_data_lake)} data lake items for document {document_id}")
        
        return jsonify({
            'success': True,
            'dataLake': document_data_lake,
            'documentId': document_id,
            'count': len(document_data_lake)
        })
        
    except Exception as e:
        logger.error(f"Error getting data lake: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/data-lake', methods=['POST'])
def save_data_lake_endpoint():
    """Save data lake items for a specific document."""
    try:
        data = request.get_json()
        
        document_id = data.get('documentId')
        window_id = data.get('windowId', 'default')
        session_id = data.get('session_id', 'default')
        data_lake_items = data.get('dataLake', [])
        
        if not document_id:
            return jsonify({
                'success': False,
                'error': 'Missing documentId in request'
            }), 400
        
        # Store data lake items for this document
        data_lake_storage[document_id] = data_lake_items
        
        # Persist to file
        save_data_lake(data_lake_storage)
        
        logger.info(f"🗂️ Saved {len(data_lake_items)} data lake items for document {document_id}")
        
        return jsonify({
            'success': True,
            'message': f'Data lake saved for document {document_id}',
            'documentId': document_id,
            'count': len(data_lake_items)
        })
        
    except Exception as e:
        logger.error(f"Error saving data lake: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/variables', methods=['GET'])
def get_variables():
    """Get variables for a specific document."""
    try:
        document_id = request.args.get('documentId')
        window_id = request.args.get('windowId', 'default')
        session_id = request.args.get('session_id', 'default')
        
        if not document_id:
            return jsonify({
                'success': False,
                'error': 'Missing documentId parameter'
            }), 400
        
        # Get variables for this document
        document_variables = variables_storage.get(document_id, {})
        
        logger.info(f"📊 Returning {len(document_variables)} variables for document {document_id}")
        
        return jsonify({
            'success': True,
            'variables': document_variables,
            'documentId': document_id,
            'count': len(document_variables)
        })
        
    except Exception as e:
        logger.error(f"Error getting variables: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/variables', methods=['POST'])
def save_variables_endpoint():
    """Save variables for a specific document."""
    try:
        data = request.get_json()
        
        document_id = data.get('documentId')
        window_id = data.get('windowId', 'default')
        session_id = data.get('session_id', 'default')
        variables_data = data.get('variables', {})
        
        if not document_id:
            return jsonify({
                'success': False,
                'error': 'Missing documentId in request'
            }), 400
        
        # Store variables for this document
        variables_storage[document_id] = variables_data
        
        # Persist to file
        save_variables(variables_storage)
        
        logger.info(f"📊 Saved {len(variables_data)} variables for document {document_id}")
        
        return jsonify({
            'success': True,
            'message': f'Variables saved for document {document_id}',
            'documentId': document_id,
            'count': len(variables_data)
        })
        
    except Exception as e:
        logger.error(f"Error saving variables: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/suggest-variable', methods=['POST'])
def suggest_variable():
    """Get LLM-powered variable suggestions based on selected text and template context."""
    try:
        data = request.get_json()
        
        # Extract request data
        template_content = data.get('template_content', '')
        selected_text = data.get('selected_text', '')
        existing_variables = data.get('existing_variables', {})
        document_id = data.get('document_id', 'default')
        
        # Validate required fields
        if not selected_text:
            return jsonify({
                'success': False,
                'error': 'Selected text is required'
            }), 400
        
        # Check if LLM client is available
        if client is None:
            return jsonify({
                'success': False,
                'error': 'AI service not available (API key not configured)',
                'fallback_suggestion': {
                    'name': 'manual_variable',
                    'description': 'Please manually configure this variable',
                    'type': 'text',
                    'format': '',
                    'confidence': 0.1,
                    'value_to_replace': selected_text,
                    'static_prefix': '',
                    'static_suffix': ''
                }
            })
        
        # Create structured prompt for LLM
        existing_vars_text = ""
        if existing_variables:
            existing_vars_text = f"\nExisting variables in template:\n"
            for var_name, var_info in existing_variables.items():
                existing_vars_text += f"- {var_name}: {var_info.get('description', 'No description')} ({var_info.get('type', 'unknown')})\n"
        
        prompt = f"""You are an AI assistant helping to create template variables. Based on the selected text and template context, suggest appropriate variable information.

TEMPLATE CONTENT:
{template_content}

SELECTED TEXT: "{selected_text}"
{existing_vars_text}

IMPORTANT: The user selected the entire text "{selected_text}", but they likely want to keep descriptive labels and only replace the actual data values with variables.

Your task:
1. Analyze the selected text to identify what part should become a variable (usually numbers, amounts, dates, etc.)
2. Identify what parts should remain as static text (usually labels, descriptions, prefixes)
3. Suggest appropriate variable information

Consider:
- The context within the template
- The format and content of the selected text
- Avoid naming conflicts with existing variables
- Use meaningful, business-friendly names
- Detect data types from patterns ($ for currency, % for percentage, etc.)

IMPORTANT: Respond with ONLY a JSON object in this exact format:
{{
    "name": "suggested_variable_name",
    "description": "Clear description of what this variable represents",
    "type": "currency|number|percentage|date|text",
    "format": "format_string_if_applicable",
    "confidence": 0.95,
    "reasoning": "Brief explanation of why these suggestions were made",
    "value_to_replace": "the exact part that should become the variable",
    "static_prefix": "text that should remain before the variable (can be empty)",
    "static_suffix": "text that should remain after the variable (can be empty)"
}}

Requirements:
1. "name" must be valid variable name (letters, numbers, underscores only, start with letter)
2. "description" should be business-friendly and clear
3. "type" must be one of: currency, number, percentage, date, text
4. "format" should be appropriate for the type (can be empty)
5. "confidence" should be between 0.0 and 1.0
6. "value_to_replace" should be the exact substring that will become the variable
7. "static_prefix" + "value_to_replace" + "static_suffix" should equal the original selected text
8. Return ONLY the JSON object, no other text

JSON:"""
        
        try:
            # Call LLM for suggestion
            if hasattr(client, 'chat'):
                # OpenAI client
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.3,
                    max_tokens=400
                )
                suggestion_text = response.choices[0].message.content.strip()
            else:
                # Together client
                response = client.chat.completions.create(
                    model="Qwen/Qwen2.5-Coder-32B-Instruct",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.3,
                    max_tokens=400
                )
                suggestion_text = response.choices[0].message.content.strip()
            
            print(f"Variable Suggestion Raw Response: {suggestion_text}")
            
            # Parse LLM response
            try:
                import json
                import re
                
                # Clean up the response to extract JSON
                json_match = re.search(r'\{[\s\S]*\}', suggestion_text)
                if json_match:
                    json_str = json_match.group()
                    suggestion = json.loads(json_str)
                    
                    # Validate suggestion structure
                    required_fields = ['name', 'description', 'type', 'value_to_replace']
                    if all(field in suggestion for field in required_fields):
                        # Ensure name is valid
                        import re
                        if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', suggestion['name']):
                            suggestion['name'] = re.sub(r'[^a-zA-Z0-9_]', '_', suggestion['name'])
                            if not suggestion['name'][0].isalpha() and suggestion['name'][0] != '_':
                                suggestion['name'] = 'var_' + suggestion['name']
                        
                        # Ensure type is valid
                        valid_types = ['currency', 'number', 'percentage', 'date', 'text']
                        if suggestion['type'] not in valid_types:
                            suggestion['type'] = 'text'
                        
                        # Set default values for optional fields
                        suggestion.setdefault('format', '')
                        suggestion.setdefault('confidence', 0.8)
                        suggestion.setdefault('reasoning', 'AI-generated suggestion')
                        suggestion.setdefault('static_prefix', '')
                        suggestion.setdefault('static_suffix', '')
                        
                        # Validate that the parts add up to the original text
                        reconstructed = suggestion['static_prefix'] + suggestion['value_to_replace'] + suggestion['static_suffix']
                        
                        # Normalize whitespace for comparison (handle non-breaking spaces, etc.)
                        def normalize_whitespace(text):
                            return re.sub(r'\s+', ' ', text.replace('\xa0', ' ').replace('\u00a0', ' '))
                        
                        original_normalized = normalize_whitespace(selected_text)
                        reconstructed_normalized = normalize_whitespace(reconstructed)
                        
                        if reconstructed_normalized != original_normalized:
                            print(f"Warning: Reconstructed text doesn't match original after normalization.")
                            print(f"  Original: '{selected_text}' (normalized: '{original_normalized}')")
                            print(f"  Reconstructed: '{reconstructed}' (normalized: '{reconstructed_normalized}')")
                            # Fallback: treat entire text as variable
                            suggestion['value_to_replace'] = selected_text
                            suggestion['static_prefix'] = ''
                            suggestion['static_suffix'] = ''
                        else:
                            print(f"✓ Text reconstruction successful: '{original_normalized}'")
                        
                        logger.info(f"Generated variable suggestion for '{selected_text}': {suggestion['name']} (replacing '{suggestion['value_to_replace']}')")
                        
                        return jsonify({
                            'success': True,
                            'suggestion': suggestion,
                            'analysis': {
                                'selected_text_length': len(selected_text),
                                'template_length': len(template_content),
                                'existing_variables_count': len(existing_variables),
                                'document_id': document_id
                            }
                        })
                    else:
                        raise ValueError("Missing required fields in LLM response")
                else:
                    raise ValueError("No valid JSON found in LLM response")
                    
            except (json.JSONDecodeError, ValueError) as parse_error:
                print(f"Error parsing LLM response: {parse_error}")
                print(f"Raw response: {suggestion_text}")
                
                return jsonify({
                    'success': False,
                    'warning': 'Used fallback suggestion due to LLM parsing error'
                })
                
        except Exception as llm_error:
            print(f"Error calling LLM: {llm_error}")

            return jsonify({
                'success': True,
                'suggestion': {},
                'warning': 'Used fallback suggestion due to LLM error'
            })
        
    except Exception as e:
        logger.error(f"Error in suggest_variable: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'fallback_suggestion': {
                'name': 'error_variable',
                'description': 'Error occurred while generating suggestion',
                'type': 'text',
                'format': '',
                'confidence': 0.1,
                'value_to_replace': selected_text if 'selected_text' in locals() else '',
                'static_prefix': '',
                'static_suffix': ''
            }
        }), 500


if __name__ == '__main__':
    print("Starting Python backend server...")
    app.run(host='127.0.0.1', port=5000, debug=True) 