import io
import base64
import openpyxl
import pandas as pd
import json
from bs4 import BeautifulSoup
import os

# File processing utility functions
def process_excel_file(file_content, file_name):
    """Process Excel file and extract text content."""
    try:
        # Create a BytesIO object from the content
        file_buffer = io.BytesIO(base64.b64decode(file_content))
        
        # Read Excel file
        if file_name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_buffer)
            content = f"Excel File: {file_name}\n\n"
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content += f"Sheet: {sheet_name}\n"
                content += "-" * 40 + "\n"
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                        content += row_text + "\n"
                content += "\n"
        else:
            # For .xls files, use pandas
            df_dict = pd.read_excel(file_buffer, sheet_name=None)
            content = f"Excel File: {file_name}\n\n"
            
            for sheet_name, df in df_dict.items():
                content += f"Sheet: {sheet_name}\n"
                content += "-" * 40 + "\n"
                content += df.to_string(index=False) + "\n\n"
        
        return content
        
    except Exception as e:
        return f"Error processing Excel file: {str(e)}"


def process_html_file(file_content, file_name):
    """Process HTML file and extract text content."""
    try:
        # If content is base64 encoded, decode it
        if isinstance(file_content, str) and len(file_content) > 100:
            try:
                html_content = base64.b64decode(file_content).decode('utf-8')
            except:
                html_content = file_content
        else:
            html_content = file_content
        
        # Parse HTML and extract text
        soup = BeautifulSoup(html_content, 'html.parser')
        
        content = f"HTML File: {file_name}\n\n"
        
        # Extract title if available
        title = soup.find('title')
        if title:
            content += f"Title: {title.get_text().strip()}\n\n"
        
        # Extract main content
        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()
        
        # Get text content
        text = soup.get_text()
        
        # Clean up whitespace
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        content += '\n'.join(chunk for chunk in chunks if chunk)
        
        return content
        
    except Exception as e:
        return f"Error processing HTML file: {str(e)}"

def process_pptx_file(file_path):
    """
    PPTX files are now processed client-side using PPTX2HTML JavaScript library.
    This function returns a message indicating the file should be processed in the browser.
    """
    try:
        # Get basic file info
        file_size = os.path.getsize(file_path)
        
        return {
            'success': True,
            'content': f'PPTX file ready for client-side processing. File size: {file_size} bytes. This file will be processed using the PPTX2HTML JavaScript library in your browser for better visual fidelity.',
            'file_type': 'pptx',
            'processing_method': 'client-side'
        }
    except Exception as e:
        return {
            'success': False,
            'error': f'Error accessing PPTX file: {str(e)}'
        }